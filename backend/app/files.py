"""Multi-format ingest for email extraction.

Accepts the file types people actually have lying around — plain text,
CSVs exported from CRMs, XLSX spreadsheets, raw HTML, JSON dumps,
``.eml`` / ``.mbox`` exports, and log files. The strategy is deliberately
simple: collapse the file to *text* and let :func:`extract_emails` do the
real work. This keeps the pipeline regex-based and tolerant of malformed
input, while still benefiting from per-format optimisations (e.g. only
scanning cell values in an XLSX rather than every byte of the binary).
"""

from __future__ import annotations

import csv
import io
import json
import re
from typing import Callable, Iterable

from .extractor import extract_emails

_HTML_TAG_RE = re.compile(r"<[^>]+>")
_HTML_MAILTO_RE = re.compile(r"mailto:([^\"'\s<>]+)", re.IGNORECASE)

_Handler = Callable[[bytes], str]


def _flatten(value: object, out: list[str]) -> None:
    """Recursively walk a JSON-shaped value and collect every string leaf."""
    if isinstance(value, str):
        out.append(value)
    elif isinstance(value, dict):
        for v in value.values():
            _flatten(v, out)
    elif isinstance(value, (list, tuple, set)):
        for v in value:
            _flatten(v, out)


def _decode_text(data: bytes) -> str:
    """Best-effort decode for text-ish payloads. Falls back to latin-1 so
    arbitrary bytes still produce *some* text we can regex over."""
    for encoding in ("utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _from_csv(data: bytes) -> str:
    """Concatenate every cell in a CSV/TSV with newlines so the regex can
    pick up embedded addresses regardless of which column they sit in."""
    text = _decode_text(data)
    chunks: list[str] = []
    # Sniff dialect; fall back to comma if sniffing fails.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    for row in reader:
        for cell in row:
            if cell:
                chunks.append(cell)
    return "\n".join(chunks)


def _from_xlsx(data: bytes) -> str:
    """Extract every visible string cell from an XLSX. Requires
    :mod:`openpyxl`; raises :class:`RuntimeError` if it isn't installed."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - dep is in pyproject.toml
        raise RuntimeError(
            "openpyxl is required to parse .xlsx files; add it to the backend deps."
        ) from exc

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    chunks: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                chunks.append(str(cell))
    return "\n".join(chunks)


def _from_json(data: bytes) -> str:
    text = _decode_text(data)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # Maybe NDJSON / JSONL
        lines: list[str] = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                parsed = json.loads(line)
            except json.JSONDecodeError:
                lines.append(line)
                continue
            collected: list[str] = []
            _flatten(parsed, collected)
            lines.extend(collected)
        return "\n".join(lines)
    collected2: list[str] = []
    _flatten(parsed, collected2)
    return "\n".join(collected2)


def _from_html(data: bytes) -> str:
    text = _decode_text(data)
    # Pull mailto: hrefs out of the tag attributes *before* the tag stripper
    # eats them, then strip tags but keep their text content.
    mailto_hits = _HTML_MAILTO_RE.findall(text)
    stripped = _HTML_TAG_RE.sub(" ", text)
    if mailto_hits:
        stripped = stripped + "\n" + "\n".join(mailto_hits)
    return stripped


def _from_eml(data: bytes) -> str:
    """Treat .eml/.mbox as plain text — every header (From/To/Cc/Bcc) plus
    body content can contribute addresses, and the regex handles all of
    them in one pass."""
    return _decode_text(data)


_DISPATCH: dict[str, _Handler] = {}


def _register(*exts: str) -> Callable[[_Handler], _Handler]:
    def deco(fn: _Handler) -> _Handler:
        for ext in exts:
            _DISPATCH[ext.lower()] = fn
        return fn

    return deco


@_register("txt", "log", "tsv", "md", "rtf")
def _txt(data: bytes) -> str:
    return _decode_text(data)


@_register("csv")
def _csv(data: bytes) -> str:
    return _from_csv(data)


@_register("xlsx", "xlsm", "xltx", "xltm")
def _xlsx(data: bytes) -> str:
    return _from_xlsx(data)


@_register("html", "htm", "xhtml")
def _html(data: bytes) -> str:
    return _from_html(data)


@_register("json", "jsonl", "ndjson")
def _json(data: bytes) -> str:
    return _from_json(data)


@_register("eml", "mbox", "msg")
def _eml(data: bytes) -> str:
    return _from_eml(data)


def supported_extensions() -> list[str]:
    return sorted(_DISPATCH.keys())


def file_to_text(filename: str, data: bytes) -> str:
    """Convert ``(filename, raw_bytes)`` to a flat text representation
    suitable for :func:`extract_emails`."""
    if not data:
        return ""
    name = (filename or "").lower()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""
    handler = _DISPATCH.get(ext)
    if handler is None:
        # Unknown extension — fall back to text decode + html-tag-strip
        # so HTML downloaded with a .download extension still works.
        text = _decode_text(data)
        if "<" in text and ">" in text and ("@" in text):
            text = _HTML_TAG_RE.sub(" ", text)
        return text
    return handler(data)


def extract_from_file(filename: str, data: bytes) -> list[str]:
    """End-to-end: turn an uploaded file into a deduped list of emails."""
    text = file_to_text(filename, data)
    return extract_emails(text)


def emails_from_iterable(items: Iterable[str]) -> list[str]:
    """Turn an arbitrary iterable of strings into a deduped, lower-cased
    email list. Used when a CSV is *known* to be one address per row and
    we don't want to lose ordering or noise from non-email cells."""
    seen: set[str] = set()
    out: list[str] = []
    for raw in items:
        if not raw:
            continue
        addr = raw.strip().lower()
        if not addr or addr in seen:
            continue
        seen.add(addr)
        out.append(addr)
    return out
